#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Author: Rui Ren
# Copyright (c) 2025 CONTINENTAL AUTOMOTIVE. All rights reserved.
# Description: Convert DiDi PDF trip invoices to Excel summary with statistics.
# Version: 1.0
# Date: 2025-06-05

import os
import re
import glob
import pandas as pd
import pdfplumber
from datetime import datetime
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00

def extract_table_from_pdf(pdf_path):
    """从PDF文件中提取表格数据"""
    all_rows = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # 提取表格
            tables = page.extract_tables()
            
            for table in tables:
                # 跳过表头行
                if table and len(table) > 1:
                    # 检查是否是我们需要的表格（包含序号、车型等列）
                    header = table[0]
                    if any('序号' in str(cell) for cell in header) and any('车型' in str(cell) for cell in header):
                        # 跳过表头，处理数据行
                        for row in table[1:]:
                            # 过滤掉空行
                            if row and any(cell for cell in row if cell):
                                # 确保行有足够的列
                                while len(row) < len(header):
                                    row.append("")
                                all_rows.append(row)
    
    return all_rows

def process_data(rows):
    """处理提取的数据，转换为所需的格式"""
    processed_data = []
    
    for row in rows:
        # 跳过空行或无效行
        if not row or not row[0]:
            continue
        
        try:
            # 提取日期和时间
            date_time = row[2]  # 上车时间列
            # 如果包含换行符，则删除所有换行符
            if isinstance(date_time, str) and '\n' in date_time:
                date_time = date_time.replace('\n', '')
            if date_time and isinstance(date_time, str):
                # 尝试解析日期时间格式 (05-06 20:40 周二)
                match = re.search(r'(\d{2}-\d{2})\s*(\d{2}:\d{2})', date_time)
                if match:
                    date = match.group(1)  # 例如 05-06
                    time = match.group(2)  # 例如 20:40
                else:
                    # 如果无法解析，使用原始值
                    parts = date_time.split()
                    if len(parts) >= 2:
                        date = parts[0]
                        time = parts[1]
                    else:
                        date = date_time
                        time = ""
            else:
                date = ""
                time = ""
            
            # 提取起点和终点
            from_location = row[4] if len(row) > 4 else ""  # 起点列
            to_location = row[5] if len(row) > 5 else ""    # 终点列
            
            # 提取费用并转换为浮点数
            fee = row[7] if len(row) > 7 else ""  # 金额列
            if isinstance(fee, str):
                # 移除非数字字符
                fee_str = re.sub(r'[^\d.]', '', fee)
                # 转换为浮点数
                try:
                    fee = float(fee_str) if fee_str else 0.0
                except ValueError:
                    fee = 0.0
            
            # 创建一行数据
            processed_row = {
                'Date': date,
                'Time': time,
                'From': from_location,
                'To': to_location,
                'Taxi Fee': fee,  # 现在是浮点数
                # 添加排序用的日期时间字段
                'SortDate': convert_to_sortable_date(date, time)
            }
            
            processed_data.append(processed_row)
        
        except Exception as e:
            print("Error processing row: {}".format(row))
            print("Error message: {}".format(str(e)))
            continue
    
    return processed_data

def convert_to_sortable_date(date_str, time_str):
    """将日期和时间转换为可排序的格式"""
    try:
        # 假设日期格式为 MM-DD
        if date_str and '-' in date_str:
            month, day = date_str.split('-')
            # 使用当前年份，因为滴滴出行行程单通常不包含年份
            year = datetime.now().year
            
            # 如果月份大于当前月份，可能是去年的记录
            current_month = datetime.now().month
            if int(month) > current_month:
                year -= 1
                
            date_part = "{}-{}-{}".format(year, month, day)
            
            # 如果有时间，添加时间部分
            if time_str:
                return "{}T{}:00".format(date_part, time_str)
            else:
                return "{}T00:00:00".format(date_part)
        else:
            # 如果无法解析日期，返回一个很早的日期作为默认值
            return "1900-01-01T00:00:00"
    except Exception:
        # 出错时返回一个很早的日期
        return "1900-01-01T00:00:00"

def save_to_excel(data, output_file):
    """将处理后的数据保存到Excel文件"""
    if not data:
        print("No data to save.")
        return False
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 按日期和时间排序
    if 'SortDate' in df.columns:
        df = df.sort_values(by='SortDate')
        # 删除排序用的列
        df = df.drop(columns=['SortDate'])
    
    # 保存到Excel，不包含索引
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    # 获取工作簿和工作表
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # 设置列宽
    for idx, col in enumerate(['A', 'B', 'C', 'D', 'E']):
        if col == 'C' or col == 'D':  # 起点和终点列宽度更大
            worksheet.column_dimensions[col].width = 40
        elif col == 'A' or col == 'B':  # 日期和时间列
            worksheet.column_dimensions[col].width = 15
        else:  # 费用列
            worksheet.column_dimensions[col].width = 10
    
    # 设置金额列为数字格式
    fee_column = 5  # E列是金额列
    for row in range(2, len(df) + 2):  # 从第2行开始（跳过表头）
        cell = worksheet.cell(row=row, column=fee_column)
        cell.number_format = FORMAT_NUMBER_00  # 设置为两位小数的数字格式
    
    # 添加总计行
    row_count = len(df) + 2  # 表头行 + 数据行数 + 1（因为Excel行从1开始）
    
    # 添加"总计"标签
    total_label_cell = worksheet.cell(row=row_count, column=1)  # A列
    total_label_cell.value = "Total"
    total_label_cell.font = Font(bold=True)
    
    # 添加求和公式
    fee_column_letter = 'E'  # 金额列
    first_data_row = 2  # 第一行数据（跳过表头）
    last_data_row = row_count - 1  # 最后一行数据
    
    sum_formula = "=SUM({}{}:{}{})".format(
        fee_column_letter, first_data_row,
        fee_column_letter, last_data_row
    )
    
    sum_cell = worksheet.cell(row=row_count, column=5)  # E列
    sum_cell.value = sum_formula
    sum_cell.font = Font(bold=True)
    sum_cell.number_format = FORMAT_NUMBER_00  # 设置数字格式为两位小数
    
    # 设置边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 为总计行添加边框
    total_label_cell.border = thin_border
    sum_cell.border = thin_border
    
    # 保存Excel文件
    writer.close()
    
    print("Data saved to {}".format(output_file))
    return True

def main():
    # 获取当前目录下所有PDF文件
    pdf_files = glob.glob("*.pdf")
    
    if not pdf_files:
        print("No PDF files found in the current directory.")
        return
    
    all_data = []
    
    # 处理每个PDF文件
    for pdf_file in pdf_files:
        print("Processing file: {}".format(pdf_file))
        rows = extract_table_from_pdf(pdf_file)
        data = process_data(rows)
        all_data.extend(data)
    
    # 生成输出文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = "滴滴出行行程报销单_{}.xlsx".format(timestamp)
    
    # 保存数据
    save_to_excel(all_data, output_file)

if __name__ == "__main__":
    main()